# %%
from datetime import datetime
import difflib
import re
import openpyxl
import pandas as pd
from pandas import DataFrame
from openpyxl.styles import PatternFill
from IPython.display import display

def normalize_string(name):
            name = re.sub(r'\s+', ' ', name).strip().lower()
            name = re.sub(r'[^a-z\s]', '', name)
            return name

def save_attendance(date:str,gl:DataFrame,al:DataFrame,gl_path:str,gl_sheet_name:str):
    display(al)
    display(gl)
    students_with_signatures = []
    dateObject =  datetime.strptime(date, "%Y-%m-%d")
    matching_indices_gl = gl[gl.apply(lambda row: row.astype(str).str.contains('Name|Vorname|DO-Ma.Nr.|GM-Ma.Nr.|Termine|Anzahl der Termine', na=False).any(), axis=1)].index[0]
    matching_indices_al = al[al.apply(lambda row: row.astype(str).str.contains('Nr.|Name|Vorname|Druckschrift!|Hochschule|Unterschrift|Matr.-Nr. oder FH-zugehörigkeit', na=False).any(), axis=1)].index[0]
    row_with_dates = gl.iloc[matching_indices_gl+1,:]
    date_column = row_with_dates[row_with_dates.astype(str).eq(str(dateObject))].index
    date_column_index = gl.columns.get_loc(date_column[0])
    al =  al.iloc[matching_indices_al:].reset_index(drop=True)

    signature_col = None

    for row in al.itertuples():
            index = row.Index
            if index == 0:
                signature_col = 4 if difflib.SequenceMatcher(None, normalize_string(str(row[4])), normalize_string('Unterschrift')).ratio() > 0.7 else 3
            if (signature_col == 4 and (index == 0 or index == al.tail(1).index[0])) or (signature_col == 3 and index == 0):
                continue
            if pd.notna(row[2]) and pd.notna(row[signature_col]):
                students_with_signatures.append(row[2])

    if students_with_signatures:
        for student in students_with_signatures:
            
            gl = pd.read_excel(gl_path, sheet_name=gl_sheet_name)
            xlc = gl.copy()
            xlc.columns = xlc.iloc[matching_indices_gl]
            xlc = xlc.iloc[matching_indices_gl+1:].reset_index(drop=True)
            mask = xlc.iloc[:, [0, 1]].isna().all(axis=1)
            empty_row_positions = xlc.index[mask].tolist()

            for index,row in xlc.iterrows():
                if index > 0:
                        
                        student_name = f"{row['Name']}, {row['Vorname']}"
                        name_normalized = normalize_string(student)
                        name_normalized_gl = normalize_string(student_name)
                        name_match = difflib.SequenceMatcher(None, name_normalized, name_normalized_gl).ratio()
                        matching_student_row = matching_indices_gl+index

                        if name_match >=0.8:
                                print('Student gefunden:',' Gesamtliste:',student_name,'Anwesenheitsliste:',student)
                                workbook = openpyxl.load_workbook(gl_path)
                                sheet = workbook[gl_sheet_name]
                                current_value = sheet.cell(row=matching_student_row+matching_indices_gl, column=date_column_index+1).value
                                if not current_value:
                                        sheet.cell(row=matching_student_row+matching_indices_gl, column=date_column_index+1, value=str('x'))
                                        workbook.save(gl_path)
                                        print('Anwesenheit eingetragen für:', student)
                                        break
                                else: 
                                        print('Anwesenheit ist bereits eingetragen für:', student)
                                        workbook.close()
                                        break
                if index == empty_row_positions[1]:
                      
                      workbook = openpyxl.load_workbook(gl_path)
                      sheet = workbook[gl_sheet_name]
                      row = empty_row_positions[1]+matching_indices_gl*2
                      splitted_name = student.split()

                      if len(empty_row_positions)==2:
                        sheet.insert_rows(row+1)
                        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell1 = sheet.cell(row=row, column=1, value=str(splitted_name[0]))
                        cell1.fill = fill
                        if len(splitted_name) >1:
                                surname = ' '.join([splitted_name[i] for i in range(1, len(splitted_name))])
                                cell2 = sheet.cell(row=row, column=2, value=str(surname))
                                cell2.fill = fill
                        sheet.cell(row=row, column=date_column_index+1, value=str('x'))
                        row = sheet[int(empty_row_positions[1]+matching_indices_gl*2)]
                        row_values = [cell.value for cell in row]
                        print(row_values)
                        workbook.save(gl_path)
                        workbook.close()

                      else:
                        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell1 = sheet.cell(row=row, column=1, value=str(splitted_name[0]))
                        cell1.fill = fill
                        if len(splitted_name) >1:
                                surname = ' '.join([splitted_name[i] for i in range(1, len(splitted_name))])
                                cell2 = sheet.cell(row=row, column=2, value=str(surname))
                                cell2.fill = fill
                        sheet.cell(row=row, column=date_column_index+1, value=str('x'))
                        row = sheet[int(empty_row_positions[1]+matching_indices_gl*2)]
                        row_values = [cell.value for cell in row]
                        print(row_values)
                        workbook.save(gl_path)
                        workbook.close()

def save_online_attendance(date:str,gl:pd.DataFrame,al:pd.DataFrame,gl_path:str,gl_sheet_name:str):
    display(al)
    display(gl)
    attendees = []
    dateObject =  datetime.strptime(date, "%Y-%m-%d")
    matching_indices_gl = gl[gl.apply(lambda row: row.astype(str).str.contains('Name|Vorname|DO-Ma.Nr.|GM-Ma.Nr.|Termine|Anzahl der Termine', na=False).any(), axis=1)].index[0]
    row_with_dates = gl.iloc[matching_indices_gl+1,:]
    date_column = row_with_dates[row_with_dates.astype(str).eq(str(dateObject))].index
    date_column_index = gl.columns.get_loc(date_column[0])

    for index,row in al.iterrows():
            if not pd.isna(row.iloc[0]):
                  attendees.append(str(row.iloc[0]))
        
    if attendees:
        for student in attendees:
            gl = pd.read_excel(gl_path, sheet_name=gl_sheet_name)
            xlc = gl.copy()
            xlc.columns = xlc.iloc[matching_indices_gl]
            xlc = xlc.iloc[matching_indices_gl+1:].reset_index(drop=True)
            print(student, ' wird ausgewertet...')
            print('date column:',date_column_index)
            mask = xlc.iloc[:, [0, 1]].isna().all(axis=1)
            #display(mask)
            empty_row_positions = xlc.index[mask].tolist()
            print('Freie Zeilen:',empty_row_positions)

            for index,row in xlc.iterrows():
                if index > 0:

                        student_name = f"{row['Name']}, {row['Vorname']}"
                        name_normalized = normalize_string(student)
                        name_normalized_gl = normalize_string(student_name)
                        #print('Student from list', name_normalized,'Student from gl:',name_normalized_gl)
                        name_match = difflib.SequenceMatcher(None, name_normalized, name_normalized_gl).ratio()
                        matching_student_row = matching_indices_gl+index
                        #print('GL:',matching_indices_gl,'index:',index,'sum:',matching_student_row)
                        
                        if name_match >=0.8:
                                print('Student gefunden:',' Gesamtliste:',student_name,'Anwesenheitsliste:',student)
                                workbook = openpyxl.load_workbook(gl_path)
                                sheet = workbook[gl_sheet_name]
                                current_value = sheet.cell(row=matching_student_row+matching_indices_gl, column=date_column_index+1).value
                                print('row student:',matching_student_row+matching_indices_gl)
                                row = sheet[int(matching_student_row+matching_indices_gl)]
                                row_values = [cell.value for cell in row]
                                if not current_value:
                                        sheet.cell(row=matching_student_row+matching_indices_gl, column=date_column_index+1, value=str('x'))
                                        workbook.save(gl_path)
                                        print('Anwesenheit eingetragen für:', student)
                                        break
                                else:
                                        print('Anwesenheit ist bereits eingetragen für:', student)
                                        workbook.close()
                                        break       
                if index == empty_row_positions[1]:
                      
                      workbook = openpyxl.load_workbook(gl_path)
                      sheet = workbook[gl_sheet_name]
                      row = empty_row_positions[1]+matching_indices_gl*2
                      splitted_name = student.split()

                      if len(empty_row_positions)==2:
                        sheet.insert_rows(row+1)
                        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell1 = sheet.cell(row=row, column=1, value=str(splitted_name[0]))
                        cell1.fill = fill
                        if len(splitted_name) >1:
                                surname = ' '.join([splitted_name[i] for i in range(1, len(splitted_name))])
                                cell2 = sheet.cell(row=row, column=2, value=str(surname))
                                cell2.fill = fill
                        sheet.cell(row=row, column=date_column_index+1, value=str('x'))
                        row = sheet[int(empty_row_positions[1]+matching_indices_gl*2)]
                        row_values = [cell.value for cell in row]
                        print(row_values)
                        workbook.save(gl_path)
                        workbook.close()

                      else:
                        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell1 = sheet.cell(row=row, column=1, value=str(splitted_name[0]))
                        cell1.fill = fill
                        if len(splitted_name) >1:
                                surname = ' '.join([splitted_name[i] for i in range(1, len(splitted_name))])
                                cell2 = sheet.cell(row=row, column=2, value=str(surname))
                                cell2.fill = fill
                        sheet.cell(row=row, column=date_column_index+1, value=str('x'))
                        row = sheet[int(empty_row_positions[1]+matching_indices_gl*2)]
                        row_values = [cell.value for cell in row]
                        print(row_values)
                        workbook.save(gl_path)
                        workbook.close()


